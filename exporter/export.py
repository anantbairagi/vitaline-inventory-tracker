"""Shared workbook-to-dashboard JSON export logic."""

from __future__ import annotations

import datetime
import gc
import json
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook

SHEET_PAR = "PAR_Output"
SHEET_ACTIVATION = "Office_Activation"
SHEET_CONTROLS = "Live_Controls"
REQUIRED_SHEETS = (SHEET_PAR, SHEET_ACTIVATION, SHEET_CONTROLS)
ROLLING_WEEKS = 12

# 0-based column indices when reading PAR_Output rows (cols 1-29).
PAR_IDX = {
    "office": 0,
    "include": 1,
    "vendor": 2,
    "item": 3,
    "unit_basis": 6,
    "units_per_pack": 7,
    "price_per_pack": 9,
    "par_target": 20,
    "rop": 21,
    "on_hand": 22,
    "order_packs": 24,
    "order_units": 25,
    "order_cost": 26,
    "status": 28,
}
ACTIVATION_FIRST_ROW, ACTIVATION_LAST_ROW = 5, 23

ROW_FIELD_ORDER = [
    "office",
    "vendor",
    "item",
    "unit_basis",
    "units_per_pack",
    "price_per_pack",
    "on_hand",
    "par_target",
    "rop",
    "order_packs",
    "order_units",
    "order_cost",
    "status",
]


class ExportError(Exception):
    """Raised when a workbook cannot be exported."""


def fmt_date(v: Any) -> str:
    if isinstance(v, datetime.datetime):
        return f"{v.month}/{v.day}/{str(v.year)[2:]}"
    if isinstance(v, (int, float)):
        d = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=float(v))
        return f"{d.month}/{d.day}/{str(d.year)[2:]}"
    return "" if v is None else str(v)


def snapshot_iso(v: Any) -> str:
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)):
        d = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=float(v))
        return d.strftime("%Y-%m-%d")
    return datetime.date.today().strftime("%Y-%m-%d")


def num(v: Any) -> float:
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return 0


def _validate_sheetnames(sheetnames: list[str]) -> None:
    missing = [sheet for sheet in REQUIRED_SHEETS if sheet not in sheetnames]
    if missing:
        found = ", ".join(sheetnames)
        raise ExportError(
            f"Expected sheet(s) {', '.join(missing)} not found. Found: {found}"
        )


def _read_controls(ctl) -> dict[str, str]:
    booking_start = booking_end = demand_basis = snapshot = None
    for row_idx, row in enumerate(
        ctl.iter_rows(min_row=1, max_row=8, min_col=2, max_col=2, values_only=True),
        start=1,
    ):
        val = row[0] if row else None
        if row_idx == 2:
            booking_start = val
        elif row_idx == 3:
            booking_end = val
        elif row_idx == 7:
            demand_basis = val
        elif row_idx == 8:
            snapshot = val
    return {
        "booking_start": fmt_date(booking_start),
        "booking_end": fmt_date(booking_end),
        "snapshot": fmt_date(snapshot),
        "demand_basis": str(demand_basis or ""),
        "_snapshot_raw": snapshot,
    }


def _read_activation(act) -> dict[str, str]:
    activation: dict[str, str] = {}
    for row in act.iter_rows(
        min_row=ACTIVATION_FIRST_ROW,
        max_row=ACTIVATION_LAST_ROW,
        min_col=1,
        max_col=2,
        values_only=True,
    ):
        office = row[0] if row else None
        if office not in (None, ""):
            active = row[1] if len(row) > 1 else "Y"
            activation[str(office)] = str(active or "Y")
    return activation


def _read_par_rows(par) -> list[list]:
    rows: list[list] = []
    for row in par.iter_rows(min_row=2, min_col=1, max_col=29, values_only=True):
        if not row:
            continue
        office = row[PAR_IDX["office"]]
        if office in (None, ""):
            continue
        include = row[PAR_IDX["include"]]
        if str(include) != "Y":
            continue
        rec = {
            "office": str(office),
            "vendor": str(row[PAR_IDX["vendor"]] or ""),
            "item": str(row[PAR_IDX["item"]] or ""),
            "unit_basis": str(row[PAR_IDX["unit_basis"]] or ""),
            "units_per_pack": num(row[PAR_IDX["units_per_pack"]]),
            "price_per_pack": num(row[PAR_IDX["price_per_pack"]]),
            "on_hand": num(row[PAR_IDX["on_hand"]]),
            "par_target": num(row[PAR_IDX["par_target"]]),
            "rop": num(row[PAR_IDX["rop"]]),
            "order_packs": num(row[PAR_IDX["order_packs"]]),
            "order_units": num(row[PAR_IDX["order_units"]]),
            "order_cost": num(row[PAR_IDX["order_cost"]]),
            "status": str(row[PAR_IDX["status"]] or ""),
        }
        rows.append([rec[f] for f in ROW_FIELD_ORDER])
    return rows


def build_payload(wb) -> tuple[dict, dict, list, Any]:
    """Build dashboard payload using streaming reads (read-only workbooks)."""
    ctl = wb[SHEET_CONTROLS]
    act = wb[SHEET_ACTIVATION]
    par = wb[SHEET_PAR]

    controls_raw = _read_controls(ctl)
    snapshot_raw = controls_raw.pop("_snapshot_raw")
    controls = controls_raw
    activation = _read_activation(act)
    rows = _read_par_rows(par)
    return controls, activation, rows, snapshot_raw


def export_workbook(in_path: Path, out_dir: Path) -> dict[str, Any]:
    """Export workbook to dashboard JSON files. Returns a summary dict."""
    in_path = Path(in_path)
    out_dir = Path(out_dir)
    if not in_path.exists():
        raise ExportError(f"Workbook not found: {in_path}")

    snap_dir = out_dir / "snapshots"
    snap_dir.mkdir(parents=True, exist_ok=True)

    wb = None
    try:
        wb = load_workbook(in_path, read_only=True, data_only=True)
        _validate_sheetnames(wb.sheetnames)
        controls, activation, rows, snapshot_raw = build_payload(wb)
    except BadZipFile as exc:
        raise ExportError("Invalid or corrupted .xlsx file") from exc
    finally:
        if wb is not None:
            wb.close()
        gc.collect()

    snap_key = snapshot_iso(snapshot_raw)
    generated = datetime.datetime.now().isoformat(timespec="seconds")

    payload = {
        "generated": generated,
        "week": snap_key,
        "controls": controls,
        "activation": activation,
        "rows": rows,
    }
    compact = json.dumps(payload, separators=(",", ":"))

    (out_dir / "dashboard-data.json").write_text(compact, encoding="utf-8")
    (snap_dir / f"{snap_key}.json").write_text(compact, encoding="utf-8")

    weeks: list[dict[str, str]] = []
    for f in snap_dir.glob("*.json"):
        if f.name == "manifest.json":
            continue
        key = f.stem
        try:
            d = json.loads(f.read_text(encoding="utf-8"))
            label = d.get("controls", {}).get("snapshot", key)
        except Exception:
            label = key
        weeks.append({"week": key, "label": label, "file": f"snapshots/{f.name}"})
    weeks.sort(key=lambda w: w["week"], reverse=True)
    rolling = weeks[:ROLLING_WEEKS]

    manifest = {"generated": generated, "current": snap_key, "weeks": rolling}
    (snap_dir / "manifest.json").write_text(
        json.dumps(manifest, separators=(",", ":")), encoding="utf-8"
    )

    active = sum(1 for v in activation.values() if v == "Y")
    return {
        "week": snap_key,
        "generated": generated,
        "rows": len(rows),
        "offices": len(activation),
        "active_offices": active,
        "manifest_weeks": len(rolling),
        "snapshot_label": controls.get("snapshot", snap_key),
    }
